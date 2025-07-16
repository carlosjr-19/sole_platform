import os
import polars as pl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as Img
from openpyxl.styles import numbers
from flask import current_app

def limpiar_duplicados(csv: pl.DataFrame):
    """
    Limpia los duplicados de un DataFrame de Polars y devuelve un DataFrame limpio.
    """
    # Identificar duplicados (todas las filas excepto la primera aparición de cada 'msisdn')
    duplicados = csv.filter(csv.select(pl.col("msisdn").is_duplicated()).to_series())

    # Mostrar resultados
    print("Duplicados:\n", duplicados.select("msisdn"))

    # Eliminar duplicados, dejando la primera aparición de cada 'msisdn'
    csv_sin_duplicados = csv.unique(subset="msisdn", keep="first")

    # Mostrar cuántos quedaron
    print(f"Activaciones únicas: {csv_sin_duplicados.height}")

    return duplicados, csv_sin_duplicados

def procesar_comisiones(df: pl.DataFrame, comision_sales: str, comision: str, fecha: str):
    
    # Limpiar precios
    cols_a_limpiar = ['mvno_package_price', 'reference_price']
    for col in cols_a_limpiar:
        df = df.with_columns(
            pl.col(col)
            .cast(pl.Utf8)
            .str.replace_all(r'[\$,]', '')
            .str.strip_chars()
            .cast(pl.Float64)
            .fill_null(0.0)
            .alias(col)
        )
    
    # Crear columnas
    df = df.with_columns([
        pl.lit(0.0).alias('comisión'),
        pl.lit(0.0).alias('bonificación fija $'),
        pl.lit(0.0).alias('bono $')
    ])

    porcentaje_comision = 0.2 if comision == "20" else 0.15 if comision == "15" else 0

    precios_iguales = (df.select((pl.col('mvno_package_price') == pl.col('reference_price')).all()).item())

    df = df.with_columns([
        pl.when(pl.col('mvno_package_price') == pl.col('reference_price'))
        .then(pl.col('mvno_package_price') * porcentaje_comision)
        .otherwise((pl.col('reference_price') * porcentaje_comision) + (pl.col('mvno_package_price') - pl.col('reference_price')))
        .alias('comisión'),

        pl.when(pl.col('mvno_package_price') != pl.col('reference_price'))
        .then(pl.col('reference_price') * porcentaje_comision)
        .otherwise(0.0)
        .alias('bono $'),

        pl.when(pl.col('mvno_package_price') != pl.col('reference_price'))
        .then(pl.col('mvno_package_price') - pl.col('reference_price'))
        .otherwise(0.0)
        .alias('bonificación fija $')
    ])

    df = df.with_columns([
        pl.when(pl.col('channel') != 'Sales')
        .then(pl.col('mvno_package_price') * 0.0414)
        .otherwise(0.0)
        .alias('transacción 4.14%'),

        pl.when(pl.col('channel') != 'Sales')
        .then(3.65)
        .otherwise(0.0)
        .alias('tasa fija 3.65')
    ])

    df = df.with_columns([
        pl.col('comisión').round(2),
        pl.col('transacción 4.14%').round(2),
        pl.col('tasa fija 3.65').round(2)
    ])

    df = df.with_columns(
        (pl.col('comisión') - pl.col('transacción 4.14%') - pl.col('tasa fija 3.65')).round(2).alias('comisión_total')
    )

    df = df.with_columns([
        pl.lit(fecha).alias('mes'),
        pl.lit(comision).alias('porcentaje')
    ])


    if comision_sales == "SI":
        print(df.columns)
        df = (
            df.with_columns(
                pl.when(pl.col('channel') == 'Sales').then(0.0).otherwise(pl.col('comisión_total')).alias('comisión_total')
            )
            .with_columns(
                pl.when(pl.col('channel') == 'Sales').then(0.0).otherwise(pl.col('comisión')).alias('comisión')
            )
            .with_columns(
                pl.when(pl.col('channel') == 'Sales').then(0.0).otherwise(pl.col('bono $')).alias('bono $')
            )
            .with_columns(
                pl.when(pl.col('channel') == 'Sales').then(0.0).otherwise(pl.col('bonificación fija $')).alias('bonificación fija $')
            )

        )    

    total_comisiones = df.select(pl.col('comisión_total').sum()).item()
    total_mpp = df.select(pl.col('mvno_package_price').sum()).item()
    total_bono = df.select(pl.col('comisión').sum()).item()
    total_transaccion = df.select(pl.col('transacción 4.14%').sum()).item()
    total_desc_fijo = df.select(pl.col('tasa fija 3.65').sum()).item()
    total_bono20 = df.select(pl.col('bono $').sum()).item()
    total_bono_fijo = df.select(pl.col('bonificación fija $').sum()).item()

    columnas_finales = [
        'account_name','account_email','mvno_name', 'msisdn', 'channel', 'profile_sim', 'store_name', 'user_staff_name', 'transaction_id',
        'date', 'mes', 'mvno_package_name', 'reference_price', 'mvno_package_price', 'porcentaje', 'bono $',
        'bonificación fija $', 'comisión', 'transacción 4.14%', 'tasa fija 3.65', 'comisión_total'
    ]

    # Seleccionar columnas finales desde el dataframe original procesado
    df_final = df.select(columnas_finales)
    
    schema_final = df_final.schema

    fila_total_dict = {}
    for col, dtype in schema_final.items():
        if dtype == pl.Int64:
            fila_total_dict[col] = [0]
        elif dtype == pl.Float64:
            fila_total_dict[col] = [0.0]
        else:
            fila_total_dict[col] = ['']

    if precios_iguales:
        fila_total_dict['mvno_package_name'] = ['TOTAL']
        fila_total_dict['mvno_package_price'] = [total_mpp]
        fila_total_dict['comisión'] = [total_bono]
        fila_total_dict['transacción 4.14%'] = [total_transaccion]
        fila_total_dict['tasa fija 3.65'] = [total_desc_fijo]
        fila_total_dict['comisión_total'] = [total_comisiones]
    else:
        fila_total_dict['mvno_package_name'] = ['TOTAL']
        fila_total_dict['reference_price'] = [0.0]  
        fila_total_dict['mvno_package_price'] = [total_mpp]
        fila_total_dict['bono $'] = [total_bono20]
        fila_total_dict['bonificación fija $'] = [total_bono_fijo]
        fila_total_dict['comisión'] = [total_bono]
        fila_total_dict['transacción 4.14%'] = [total_transaccion]
        fila_total_dict['tasa fija 3.65'] = [total_desc_fijo]
        fila_total_dict['comisión_total'] = [total_comisiones]

    fila_total = pl.DataFrame(fila_total_dict)

    df_final = df_final.vstack(fila_total)

    return df_final, precios_iguales

def estilos_excel(df, marca, precios_iguales, fecha):
    # Crear workbook y hoja
    wb = Workbook()
    ws = wb.active
    ws.title = f"ACTIVACIONES {marca}"

    # Insertar imagen (ajusta 'logo.png' si tu logo tiene otro nombre)
    ruta_imagen = os.path.join(current_app.config['IMAGES_FOLDER'], "logo_bitelit.png")
    logo = Img(ruta_imagen)
    logo.anchor = 'A1'  # Esquina superior derecha
    ws.add_image(logo)

    # Encabezado "ACTIVACIONES" centrado
    ws.merge_cells('B5:L5')
    ws['B5'] = f"ACTIVACIONES {marca.upper()}"
    ws['B5'].font = Font(size=16, bold=True)
    ws['B5'].alignment = Alignment(horizontal='center')

    # Añadir columna de numeración
    df.insert(0, 'N#', range(1, len(df) + 1))

    # Sobreescribir la columna con el valor formateado
    df["mvno_package_price"] = df["mvno_package_price"].replace(r'[\$,]', '', regex=True).astype(float)
    df["comisión"] = df["comisión"].replace(r'[\$,]', '', regex=True).astype(float)
    df["transacción 4.14%"] = df["transacción 4.14%"].replace(r'[\$,]', '', regex=True).astype(float)
    df["tasa fija 3.65"] = df["tasa fija 3.65"].replace(r'[\$,]', '', regex=True).astype(float)
    df["comisión_total"] = df["comisión_total"].replace(r'[\$,]', '', regex=True).astype(float)

    # Convertir tu CSV a DataFrame para exportarlo
    df_final = df

    # Insertar datos a partir de la fila 7
    for r_idx, row in enumerate(dataframe_to_rows(df_final, index=False, header=True), start=7):
        for c_idx, value in enumerate(row, start=2):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            if precios_iguales == True:
                # Convertir a negativo si es una de las columnas que deben ser negativas
                # Las columnas U y V son las 21 y 22 en Excel (ya que empiezas en columna 2)
                if c_idx in [21, 22] and r_idx > 7:  # r_idx > 7 para no afectar los encabezados
                    try:
                        if value is not None and str(value).strip() != '':  # Solo si hay un valor
                            value = -abs(float(value))
                    except (ValueError, TypeError):
                        pass  # Si no se puede convertir a número, dejamos el valor original
            else:
                # Convertir a negativo si es una de las columnas que deben ser negativas
                # Las columnas U y V son las 21 y 22 en Excel (ya que empiezas en columna 2)
                if c_idx in [21, 22] and r_idx > 7:  # r_idx > 7 para no afectar los encabezados
                    try:
                        if value is not None and str(value).strip() != '':  # Solo si hay un valor
                            value = -abs(float(value))
                    except (ValueError, TypeError):
                        pass  # Si no se puede convertir a número, dejamos el valor original
            
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            # Formato de encabezados (fila 7)
            if r_idx == 7:
                cell.font = Font(bold=True, color="000000")
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')

                # Borde negro
                thin = Side(border_style="thin", color="000000")
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

            # Bordes para datos normales
            if r_idx > 7:
                thin = Side(border_style="thin", color="000000")
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

                # Centrar valores de columna N# (columna 2 en Excel)
                if c_idx == 2:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                # Centrar valores de columna N# (columna 2 en Excel)
                if c_idx == 14:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # alinear a la derecha valores de columna  (columna 3, 15, 16, 17 y 18 en Excel)
                if c_idx == 13:
                    cell.alignment = Alignment(horizontal='right', vertical='center')

                if c_idx == 15:
                    cell.alignment = Alignment(horizontal='right', vertical='center')

                if c_idx == 16:
                    cell.alignment = Alignment(horizontal='right', vertical='center')

                if c_idx == 17:
                    cell.alignment = Alignment(horizontal='right', vertical='center')

                if c_idx == 18:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # Ajustar altura de encabezado
        ws.row_dimensions[7].height = 30

        
        # leyenda "DESCUENTO" centrado en las celdas
        ws.merge_cells('U6:V6')
        ws['U6'] = "DESCUENTO"
        ws['U6'].font = Font(size=12, bold=False)
        ws['U6'].alignment = Alignment(horizontal='center')
        ws['U6'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))
        ws['U6'].border = thin_border
        thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))
        ws['V6'].border = thin_border

        #FORMATEAR COMO MONEDA después de haber insertado todo
        for col_letter in ['O', 'P', 'R', 'S', 'T', 'U', 'v', 'W']:  # Ajusta letras según tus columnas reales
            for cell in ws[col_letter][7:]:  # desde fila 8 en adelante
                cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

        #FORMATEAR CELDAS DE COLORES AMARILLO
        for col_letter in ['U', 'V']:  # Ajusta letras según tus columnas reales
            for cell in ws[col_letter][6:]:  # desde fila 7 en adelante
                cell.fill = PatternFill(start_color="f9e79f", end_color="f9e79f", fill_type="solid")

        #FORMATEAR CELDAS DE COLORES AZUL
        for col_letter in ['R', 'S', 'T']:  # Ajusta letras según tus columnas reales
            for cell in ws[col_letter][6:]:  # desde fila 7 en adelante
                cell.fill = PatternFill(start_color="aed6f1", end_color="aed6f1", fill_type="solid")

    # Ajustar ancho de columnas (le puedes personalizar los anchos)
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Letra de columna
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Antes de exportar, añade la columna de numeración
    df.insert(0, 'N°', range(1, len(df) + 1))

    # Cálculo total de comisiones
    total_comision = df['comisión_total'].sum()

    # Crear una fila de total con None o '' en campos no numéricos
    fila_total = [''] * len(df.columns)  # llena toda la fila con ''
    fila_total[df.columns.get_loc('mes')] = df['mes'].iloc[0]  # o lo que quieras mostrar en la columna 'mes'
    fila_total[df.columns.get_loc('comisión_total')] = total_comision  # coloca total en columna correspondiente

    # Añadir la fila al final del dataframe
    df.loc[len(df)] = fila_total

    #Colocar en blanco el header del excel
    ws.merge_cells('B1:R4')
    ws.merge_cells('M5:R5')

    #ruta de carpeta de descarga
    ruta_descargas = current_app.config['DOWNLOAD_FOLDER']
    if not os.path.exists(ruta_descargas):
        print(f"Creando carpeta de descargas: {ruta_descargas}")
        os.makedirs(ruta_descargas)  # crea carpeta si no existe
    else:
        print(f"Carpeta de descargas ya existe: {ruta_descargas}")
    

    # Guardar el archivo Excel
    nombre_archivo = f"Comisiones_act_{marca}_{fecha}.xlsx"

    if os.environ.get("RAILWAY_ENVIRONMENT"):
        # Entorno de Railway (producción)
        ruta_completa = f"/tmp/{nombre_archivo}"
    else:
        # Entorno local (Windows/Linux local)
        ruta_descargas = current_app.config['DOWNLOAD_FOLDER']
        os.makedirs(ruta_descargas, exist_ok=True)
        ruta_completa = os.path.join(ruta_descargas, nombre_archivo)

    wb.save(ruta_completa)

    print(f"[DEBUG] Archivo guardado en: {ruta_completa}")

    return nombre_archivo